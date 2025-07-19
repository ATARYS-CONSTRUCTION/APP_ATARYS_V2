# =============================================================================
# SCRIPT ENCAISSEMENT TVA – ATARYS (final code client + numéro de facture)
# =============================================================================

import os
import re
import fitz
import pdfplumber
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# === PARAMÈTRES À RENSEIGNER ===
MOIS = input("Numéro du mois à traiter (ex: 03) : ").zfill(2)
ANNEE = input("Année à traiter (ex: 2025) : ")
DOSSIER_PDF = input("Chemin complet du dossier PDF à traiter : ")
nom_mois = MOIS

# === Détection HT/TTC pour Batappli via pdfplumber ===
def extraire_montants_batappli(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[-1]
        lignes = page.extract_text().splitlines()
        ht, ttc = None, None
        for line in lignes:
            upper = line.upper()
            if "TOTAL HT" in upper:
                match = re.findall(r"\d[\d\s]*,\d{2}", line)
                if match:
                    ht = float(match[-1].replace(" ", "").replace(",", "."))
            if "TOTAL TTC" in upper:
                match = re.findall(r"\d[\d\s]*,\d{2}", line)
                if match:
                    ttc = float(match[-1].replace(" ", "").replace(",", "."))
    return ht, ttc

# === Détection HT/TTC pour EBP ===
def extraire_montants_ebp(doc):
    page = doc[-1]
    lignes = page.get_text().split("\n")
    ht, ttc = None, None
    for i, line in enumerate(lignes):
        upper = line.upper().strip()
        if "TOTAL HT" in upper and ht is None:
            match = re.findall(r"\d[\d\s]*,\d{2}", upper)
            if match:
                ht = float(match[-1].replace(" ", "").replace(",", "."))
        if "TOTAL HT" in upper and ht is None:
            for j in range(1, 6):
                if i + j < len(lignes):
                    match = re.findall(r"\d[\d\s]*,\d{2}", lignes[i + j])
                    if match:
                        ht = float(match[-1].replace(" ", "").replace(",", "."))
                        break
        if "TOTAL TTC" in upper and ttc is None:
            montants = []
            for j in range(1, 11):
                if i + j < len(lignes):
                    matches = re.findall(r"\d[\d\s]*,\d{2}", lignes[i + j])
                    for m in matches:
                        try:
                            val = float(m.replace(" ", "").replace(",", "."))
                            montants.append(val)
                        except:
                            continue
            if montants:
                ttc = max(montants)
    return ht, ttc

# === Extraction du code client après le mot Facture, d'acompte ou d'avancement
def extraire_code_client(nom_fichier):
    nom_base = os.path.splitext(nom_fichier)[0].lower()
    mots = nom_base.split()
    if "facture" in mots:
        idx = mots.index("facture")
        if idx + 1 < len(mots) and mots[idx + 1] == "d'acompte":
            code_client = mots[idx + 2:]
        elif idx + 1 < len(mots) and mots[idx + 1] == "d'avancement":
            code_client = mots[idx + 2:]
        else:
            code_client = mots[idx + 1:]
    else:
        code_client = mots
    return " ".join(code_client).upper()

# === Extraction du numéro de facture : soit Numéro, soit FT Numéro
def extraire_numero_facture(nom_fichier):
    nom_base = os.path.splitext(nom_fichier)[0]
    mots = nom_base.split()
    if mots[0] == "FT" and len(mots) > 1:
        return f"{mots[0]} {mots[1]}"
    else:
        return mots[0]

# === Extraction des infos PDF
def extraire_infos_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    texte = "\n".join(page.get_text() for page in doc)
    nom_fichier = os.path.basename(pdf_path)
    type_facture = "BATAPPLI" if "CAPITAL : 36000" in texte.upper() else "EBP"

    if type_facture == "BATAPPLI":
        ht, ttc = extraire_montants_batappli(pdf_path)
    else:
        ht, ttc = extraire_montants_ebp(doc)

    numero = extraire_numero_facture(nom_fichier)
    code_client = extraire_code_client(nom_fichier)

    # Date
    date_facture = None
    for line in texte.split("\n"):
        match = re.search(r"\d{2}/\d{2}/\d{4}", line)
        if match:
            try:
                date_facture = datetime.strptime(match.group(), "%d/%m/%Y")
                break
            except:
                continue

    # Client (après "Adresse de facturation")
    client = ""
    lignes = texte.split("\n")
    for i, line in enumerate(lignes):
        if "adresse de facturation" in line.lower():
            for j in range(1, 4):
                if i + j < len(lignes):
                    candidate = lignes[i + j].strip()
                    if candidate and not any(char.isdigit() for char in candidate):
                        client = candidate
                        break
            break

    return {
        "Code client": code_client,
        "Numéro": numero,
        "Fichier": nom_fichier,
        "Date": date_facture,
        "Client": client,
        "HT": ht,
        "TVA": None,
        "TTC": ttc
    }

# === Traitement des fichiers PDF ===
if not os.path.isdir(DOSSIER_PDF):
    raise FileNotFoundError(f"📁 Dossier introuvable : {DOSSIER_PDF}")

data = []
for fichier in os.listdir(DOSSIER_PDF):
    if fichier.lower().endswith(".pdf"):
        chemin_pdf = os.path.join(DOSSIER_PDF, fichier)
        data.append(extraire_infos_pdf(chemin_pdf))

factures_df = pd.DataFrame(data)
factures_df["Date"] = pd.to_datetime(factures_df["Date"], dayfirst=True, errors='coerce')
factures_df = factures_df.sort_values("Date")

# === Création Excel ===
wb = Workbook()
ws = wb.active
ws.title = "ENCAISSEMENT TVA"

bold = Font(bold=True)
wrap = Alignment(wrap_text=True)
center = Alignment(horizontal="center")
border = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))

headers = [
    ("Code client", 20, wrap),
    ("Numéro de facture", 9, center),
    ("Nom du fichier", 30, wrap),
    ("Date", 12, center),
    ("Nom du client", 40, wrap),
    ("Montant HT", 21, center),
    ("Montant TVA", 21, center),
    ("Montant TTC", 21, center),
]

ws.append([h[0] for h in headers])
for col, (title, width, align) in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col)
    cell.font = bold
    cell.alignment = align
    cell.border = border
    ws.column_dimensions[chr(64 + col)].width = width

for i, row in factures_df.iterrows():
    ligne = [
        row["Code client"],
        row["Numéro"],
        row["Fichier"],
        row["Date"].strftime("%d/%m/%Y") if pd.notnull(row["Date"]) else "",
        row["Client"],
        row["HT"],
        "",  # formule TVA
        row["TTC"]
    ]
    ws.append(ligne)
    for col in range(1, 9):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.alignment = headers[col-1][2]
        cell.border = border
    ws.cell(row=ws.max_row, column=7).value = f"=H{ws.max_row}-F{ws.max_row}"
    ws.cell(row=ws.max_row, column=7).alignment = center
    ws.cell(row=ws.max_row, column=7).border = border

# Totaux
ws.append([""] * 4 + ["TOTAL"] + [
    f"=SUM(F2:F{ws.max_row})",
    f"=SUM(G2:G{ws.max_row})",
    f"=SUM(H2:H{ws.max_row})"
])
for col in range(5, 9):
    cell = ws.cell(row=ws.max_row, column=col)
    cell.font = bold
    cell.border = border
    cell.alignment = center

# Enregistrement
nom_fichier = f"ENCAISSEMENT {nom_mois} {ANNEE}.xlsx"
chemin_final = os.path.join(DOSSIER_PDF, nom_fichier)
wb.save(chemin_final)
print(f"\n✅ Fichier sauvegardé : {chemin_final}")
